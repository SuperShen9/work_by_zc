# -*- coding: utf-8 -*-
# author：Super.Shen
import pandas as pd
import os,openpyxl
from Daily.Func import nian,yue,ri_now
import shutil
pd.set_option('expand_frame_repr', False)  # 当列太多时不换行
pd.set_option('display.max_rows', 1000)

if len(yue) < 2:
    yue = '0' + yue
if len(ri_now) < 2:
    ri_now = '0' + ri_now

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

def active(name):
    # Create the workbook and worksheet we'll be working with
    wb = openpyxl.load_workbook(name)
    ws = wb.active
    ws['C1'] = '数据反馈'
    ws['D1'] = '数据反馈2'
     # Create a data-validation object with list validation
    dv = DataValidation(type="list", formula1='"不需要,暂时不需要,未接,设置,需要,关机,停机,空号"', allow_blank=True)
    # Optionally set a custom error message
    dv.error ='内容填写错误'
    dv.errorTitle = 'Invalid Entry'

    # Optionally set a custom prompt message
    # dv.prompt = 'Please select from the list'
    # dv.promptTitle = '请筛选'

    # Add the data-validation object to the worksheet
    ws.add_data_validation(dv)

    dv.add('C2:C5000')
    wb.save(name)

path = 'C:\\Users\Administrator\Desktop\每日资源\分配'
path_out='C:\\Users\Administrator\Desktop\每日资源'

# 进入分配资源路径
os.chdir(path)

# 读取资源
all_list = os.listdir(path)

# 判断资源是否存在
if all_list==[]:
    print('\n请放入资源，再运行代码！')
    exit()
else:
    for i in all_list:
        df = pd.read_excel(i)

        if df.shape[1] >4:
            print('\n请按照要求放入资源：增加“姓名”，“手机号”两个标题！')
        elif df.columns[1] != '手机号':
            print('\n请按照要求放入资源：“手机号”标题错误！')
        elif df.shape[0] > 4000:
            print('数据数量超过4000条，请注意资源合理分配！')
        else:
            dict1 = {0: '潘婷', 1: '王勋', 2: '刘育辉', 3: '范华东',  4: '王国平'}

            # dict1 = {0: '张永祥', 1: '张雄'}
            print('请确认分配名单：')
            print(dict1)
            ziyuan = input('\n请输入你的资源名称：')

            # 进入分配资源路径
            os.chdir(path_out)
            x = 0
            for i in range(len(dict1)):
                y = int(df.shape[0]/len(dict1))

                name = '新电销' + '-' + dict1[i] + '-' + nian + yue + ri_now + '-' + ziyuan +'.xlsx'
                if i == len(dict1)-1:
                    df.iloc[x:x + y+len(dict1), :].to_excel(name, index=False)
                    active(name)
                    print('\n{} 分配了{}条数据'.format(dict1[i], df.iloc[x:x + y+len(dict1), :].shape[0]))
                else:
                    df.iloc[x:x+y, :].to_excel(name, index=False)
                    active(name)
                    print('\n{} 分配了{}条数据'.format(dict1[i], df.iloc[x:x + y, :].shape[0]))
                x = x+y

