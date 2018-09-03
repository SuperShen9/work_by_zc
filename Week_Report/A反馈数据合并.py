# -*- coding: utf-8 -*-
# author：Super.Shen

import os, openpyxl, xlrd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED

os.chdir('D:\Super\superflag')

wbf = openpyxl.load_workbook('提取标题.xlsx')

sheet = wbf['数据反馈']

hang1 = sheet.max_row + 1

spam = {}
for row in range(2, hang1):
    flag = sheet['A' + str(row)].value.lower()
    number = sheet['B' + str(row)].value
    spam.setdefault(flag, number)

os.chdir('D:\Super\data_hebing')

k1 = 0
filepath = 'D:\Super\data_hebing'
for x, y, excels in os.walk(filepath):
    if len(excels) == 0:
        print('\n请放入数据，再启动程序')
        exit()
    baocun = openpyxl.Workbook()
    sheet = baocun.create_sheet(index=0, title='data')
    for excel in excels:
        wb = xlrd.open_workbook(excel)
        sheet1 = wb.sheets()[0]
        hang = sheet1.nrows
        lie = sheet1.ncols
        for k in range(lie):
            if str(sheet1.cell(0, k).value).lower() in spam.keys():
                kk = get_column_letter(spam[sheet1.cell(0, k).value.lower()])
                sheet[kk + '1'] = sheet1.cell(0, k).value
                sheet['A1'] = '资源来源'
                ft = Font(name='Arial', size=12, bold=True)
                ft1 = Font(name='Arial', size=12, bold=True, color=RED)
                sheet['A1'].font = ft1
                sheet[kk + '1'].font = ft
                j = 2

                a = excel.split('-')[2][:4]
                b = excel.split('-')[2][4:6]
                c = excel.split('-')[2][6:8]
                if len(excel.split('-'))>=4:
                    d = excel.split('-')[3].split('.')[0]
                else:
                    d=excel
                for i in range(1, hang):
                    sheet['A' + str(j + k1)] = d
                    sheet['B' + str(j + k1)] = str(a)+'/'+str(b)+'/'+str(c)
                    sheet['D' + str(j + k1)] = excel.split('-')[0]
                    sheet['E' + str(j + k1)] = excel.split('-')[1]
                    sheet[kk + str(j + k1)] = sheet1.cell(i, k).value
                    j += 1
        k1 += hang - 1
sheet.freeze_panes = 'A2'

baocun.save('C:\\Users\Administrator\Desktop\Feedback_Data.xlsx')
# exit()

import pandas as pd

df = pd.read_excel('C:\\Users\Administrator\Desktop\Feedback_Data.xlsx')

df_s = pd.read_excel('D:\Super\superflag\统计.xlsx')

# # 测试作用
# df['Unnamed: 1'] = df['Unnamed: 1'].apply(lambda x: pd.to_datetime(x))
# df = pd.merge(left=df, right=df_s, on=['数据反馈'], how='left')
# df.rename(columns=({'Unnamed: 1': '资源发送时间', 'Unnamed: 3': '部门', 'Unnamed: 4': '姓名'}), inplace=True)
# df.to_excel('C:\\Users\Administrator\Desktop\Feedback_Data.xlsx', index=False)
# print('\n反馈列没有问题\n')
# for col in df.columns[:-1]:
#     print(df[col].count())
# exit()

if df['手机号'].count() == df['数据反馈'].count():
    df['Unnamed: 1'] = df['Unnamed: 1'].apply(lambda x: pd.to_datetime(x))
    df = pd.merge(left=df, right=df_s, on=['数据反馈'], how='left')
    df.rename(columns=({'Unnamed: 1': '资源发送时间', 'Unnamed: 3': '部门', 'Unnamed: 4': '姓名'}), inplace=True)
    df.to_excel('C:\\Users\Administrator\Desktop\Feedback_Data.xlsx', index=False)
    print('\n反馈列没有问题\n')
    for col in df.columns[:-1]:
        print(df[col].count())

else:
    print('\n程序已删除空白列:{}条，请核对！'.format(df['数据反馈'].count()-df['手机号'].count()))
    df = df[df['数据反馈'].notnull()]
    # print(df)
    # exit()
    df['Unnamed: 1'] = df['Unnamed: 1'].apply(lambda x: pd.to_datetime(x))
    df = pd.merge(left=df, right=df_s, on=['数据反馈'], how='left')
    df.rename(columns=({'Unnamed: 1': '资源发送时间', 'Unnamed: 3': '部门', 'Unnamed: 4': '姓名'}), inplace=True)
    df.to_excel('C:\\Users\Administrator\Desktop\Feedback_Data.xlsx', index=False)

